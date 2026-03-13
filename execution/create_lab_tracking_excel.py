import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def gerar_planilha():
    filename = "Controle_Laboratorio_Odontologia.xlsx"
    
    # Cria uma nova pasta de trabalho
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestão de Trabalhos"
    
    # Cabeçalhos
    headers = [
        "Status Atual",
        "ID Trabalho / OS",
        "Paciente",
        "Dentista",
        "Laboratório",
        "Descrição do Serviço",
        "1º Envio (Saída)",
        "Previsão Retorno 1",
        "Retorno 1 (Entrada/Prova)",
        "2º Envio (Saída pra Ajuste)",
        "Previsão Retorno Final",
        "Retorno Final (Entrada)",
        "Observações / Cor"
    ]
    
    ws.append(headers)
    
    # Formatação dos Cabeçalhos
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"), 
        right=Side(style='thin', color="BFBFBF"), 
        top=Side(style='thin', color="BFBFBF"), 
        bottom=Side(style='thin', color="BFBFBF")
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Ajusta a largura das colunas baseada no cabeçalho
        ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 4
        
    ws.column_dimensions['A'].width = 25 # Status
    ws.column_dimensions['C'].width = 30 # Paciente
    ws.column_dimensions['D'].width = 20 # Dentista
    ws.column_dimensions['E'].width = 20 # Laboratório
    ws.column_dimensions['F'].width = 35 # Serviço
    ws.column_dimensions['M'].width = 40 # Obs
        
    # Validadores de Dados (Listas Suspensas)
    
    # Validador de Dentistas
    dv_dentista = DataValidation(type="list", formula1='"Rafael Nunes,Edgardo Julio"', allow_blank=True)
    ws.add_data_validation(dv_dentista)
    dv_dentista.add('D2:D1000')
    
    # Validador de Laboratorios
    dv_lab = DataValidation(type="list", formula1='"Lab 1,Lab 2,Lab 3,Lab 4,Lab 5,Lab 6"', allow_blank=True)
    ws.add_data_validation(dv_lab)
    dv_lab.add('E2:E1000')
    
    # Validador de Status
    dv_status = DataValidation(type="list", formula1='"No Laboratório (1º Envio),Disponível para Prova,Em Prova no Consultório,Reenviado (Ajuste/Finalização),Aguardando Instalação,Finalizado/Instalado,Cancelado"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('A2:A1000')

    # Dados de Exemplo
    example_data_1 = ["No Laboratório (1º Envio)", "OS-001", "João Silva", "Rafael Nunes", "Lab 1", "Prótese Protocolo Superior", "06/03/2026", "13/03/2026", "", "", "", "", "Cor A2, dente 11 com leve giro"]
    example_data_2 = ["Reenviado (Ajuste/Finalização)", "OS-002", "Maria Oliveira", "Edgardo Julio", "Lab 4", "Coroa E-max Dente 21", "28/02/2026", "04/03/2026", "03/03/2026", "05/03/2026", "10/03/2026", "", "Solicitado ajuste no contato mesial"]
    
    ws.append(example_data_1)
    ws.append(example_data_2)

    # Formatar as células de dados
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if cell.column not in [3, 6, 13] else "left", vertical="center")

    # Congelar a primeira linha para rolar a tela mantendo os cabeçalhos
    ws.freeze_panes = "A2"

    # Salva o arquivo localmente
    wb.save(filename)
    print(f"Arquivo gerado em: {filename}")

if __name__ == "__main__":
    gerar_planilha()
